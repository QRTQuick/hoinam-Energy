from __future__ import annotations

import io
import logging
import traceback
import sys
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path

from flask import Flask, Response, g, jsonify, request, send_from_directory
from flask_cors import CORS
from sqlalchemy import desc, func
from sqlalchemy.orm import joinedload

from .config import get_settings
from .database import check_database_url, close_session, get_session, init_database
from .emailer import (
    build_order_notification_message,
    build_order_approved_message,
    build_subscription_confirmation_message,
    build_new_post_notification_message,
    build_feedback_notification_message,
    build_feedback_acknowledgement_message,
    send_message_via_smtp,
    smtp_is_configured,
)
from email.message import EmailMessage
from .firebase_auth import verify_id_token

# Configure logging
logger = logging.getLogger(__name__)
from .inventory import parse_stock_inventory
from .models import (
    BlogPost,
    BlogSubscriber,
    Coupon,
    DocumentRecord,
    Feedback,
    Installation,
    InventoryMovement,
    JobListing,
    Order,
    Payment,
    Product,
    User,
)
from .seed import seed_products, seed_coupons, seed_jobs
from .services import (
    apply_product_payload,
    calculate_order_items,
    flag_duplicate_full_name_users,
    sync_user_from_claims,
    StockError,
)
from .stores import get_all_stores, get_store_by_slug
from .utils import (
    generate_order_number,
    generate_payment_reference,
    generate_verification_code,
    generate_unsubscribe_token,
    resolve_product_image_url,
    slugify,
    to_decimal,
)


class ApiError(Exception):
    def __init__(self, message: str, status_code: int = 400):
        self.message = message
        self.status_code = status_code
        super().__init__(message)


def create_app() -> Flask:
    settings = get_settings()
    app = Flask(__name__)
    project_root = Path(__file__).resolve().parents[1]
    CORS(
        app,
        resources={r"/api/*": {"origins": settings.cors_origins or "*"}},
        supports_credentials=True,
    )

    # ── Database initialisation ───────────────────────────────────────────────
    # Run once at app-creation time so the first real request isn't delayed by
    # table creation, schema migrations, and product seeding.
    try:
        check_database_url()
        init_database()
        _seed_session = get_session()
        try:
            seed_products(_seed_session)
            seed_coupons(_seed_session)
            seed_jobs(_seed_session)
            _seed_session.commit()
        except Exception:
            _seed_session.rollback()
            raise
        finally:
            close_session()
    except Exception as _init_err:
        app.logger.warning("DB init at startup failed (%s) — will retry on first request.", _init_err)

    _db_initialized = True  # mark done so before_request skips it

    def route_needs_database(path: str) -> bool:
        if not path.startswith("/api/"):
            return False
        database_optional_paths = {
            "/api/health",
            "/api/season",
            "/api/stores",
            "/api/payment-options",
        }
        if path in database_optional_paths:
            return False
        if path.startswith("/api/stores/"):
            return False
        return True

    def ensure_db_initialized():
        # No-op — init already ran above. Kept for safety in case startup failed.
        pass

    @app.before_request
    def attach_session():
        if route_needs_database(request.path):
            ensure_db_initialized()
            g.db = get_session()

    @app.teardown_request
    def teardown_request(exc):
        if hasattr(g, "db"):
            if exc is not None:
                g.db.rollback()
            close_session()

    @app.errorhandler(ApiError)
    def handle_api_error(error):
        logger.warning(f"API Error [{error.status_code}]: {error.message}")
        return jsonify({"success": False, "message": error.message}), error.status_code

    @app.errorhandler(Exception)
    def handle_unexpected_error(error):
        if request.path.startswith("/api/"):
            # Log full exception details
            logger.error(
                f"Unhandled API Exception [{request.method} {request.path}]: {type(error).__name__}: {str(error)}",
                exc_info=True,
                extra={
                    "method": request.method,
                    "path": request.path,
                    "remote_addr": request.remote_addr,
                    "error_type": type(error).__name__,
                }
            )
            # Return user-friendly error message
            return jsonify(
                {
                    "success": False,
                    "message": "An unexpected error occurred. Please try again later.",
                    "error_id": getattr(error, "error_id", None),  # For support debugging
                }
            ), 500
        raise error

    def db_session():
        return g.db

    def json_success(data=None, status_code: int = 200, message: str | None = None):
        payload = {"success": True}
        if message:
            payload["message"] = message
        if data is not None:
            payload["data"] = data
        return jsonify(payload), status_code

    def authenticate(*, required: bool = True, admin: bool = False):
        auth_header = request.headers.get("Authorization", "").strip()
        token = None
        if auth_header.lower().startswith("bearer "):
            token = auth_header.split(" ", 1)[1].strip()
        elif request.is_json:
            token = (request.get_json(silent=True) or {}).get("idToken")

        if not token:
            if required:
                logger.debug(f"Authentication required but no token provided from {request.remote_addr}")
                raise ApiError("Authentication token is required.", 401)
            return None

        try:
            claims = verify_id_token(token)
            logger.debug(f"Firebase token verified for user: {claims.get('email')}")
        except Exception as exc:
            logger.warning(f"Firebase token verification failed: {type(exc).__name__}: {str(exc)}")
            raise ApiError(f"Unable to verify authentication token. Please sign in again.", 401) from exc

        try:
            user = sync_user_from_claims(db_session(), claims)
            logger.debug(f"User synced: {user.email} (ID: {user.id})")
        except ValueError as exc:
            logger.warning(f"User sync failed: {str(exc)}")
            raise ApiError(str(exc), 409) from exc
        except Exception as exc:
            logger.error(f"Unexpected error during user sync: {type(exc).__name__}: {str(exc)}", exc_info=True)
            raise ApiError("Failed to sync user session. Please sign in again.", 500) from exc
        
        g.current_user = user

        if admin and user.role != "admin":
            logger.warning(f"Admin action attempted by non-admin user: {user.email}")
            raise ApiError("Admin access is required.", 403)

        return user

    def serialize_installation(installation: Installation) -> dict:
        payload = installation.to_dict()
        payload["user"] = installation.user.to_dict() if installation.user else None
        payload["product"] = (
            installation.product.to_dict() if installation.product else None
        )
        return payload

    def parse_inventory_file(file_storage) -> list[dict]:
        rows = parse_stock_inventory(file_storage.stream)
        if not rows:
            raise ApiError(
                "The uploaded workbook is empty or does not match the stock inventory layout.",
                400,
            )
        return rows

    def find_existing_inventory_product(
        row: dict, product_sku: str, product_slug: str, legacy_slug: str | None
    ):
        reference_text = (row.get("reference") or "").strip()
        session = db_session()

        exact_conditions = [
            Product.sku == product_sku,
            Product.slug == product_slug,
            Product.name == row["name"],
        ]
        if legacy_slug:
            exact_conditions.append(Product.slug == legacy_slug)

        for condition in exact_conditions:
            product = session.query(Product).filter(condition).first()
            if product is not None:
                return product

        if legacy_slug:
            product = (
                session.query(Product)
                .filter(Product.slug.endswith(f"-{legacy_slug}"))
                .first()
            )
            if product is not None:
                return product

        if reference_text:
            product = (
                session.query(Product)
                .filter(Product.name.endswith(f" {reference_text}"))
                .first()
            )
            if product is not None:
                return product

        return None

    def _payload_bool(payload: dict, key: str, default: bool = False) -> bool:
        value = payload.get(key, default)
        if isinstance(value, str):
            return value.strip().lower() in {"1", "true", "yes", "on"}
        return bool(value)

    def _payload_list(payload: dict, key: str) -> list[str]:
        value = payload.get(key) or []
        if isinstance(value, str):
            return [
                item.strip()
                for item in value.replace("\r", "\n").replace(",", "\n").split("\n")
                if item.strip()
            ]
        if isinstance(value, list):
            return [str(item).strip() for item in value if str(item).strip()]
        return []

    def normalize_job_payload(payload: dict, *, existing: JobListing | None = None) -> dict:
        title = (payload.get("title") or existing.title if existing else payload.get("title") or "").strip()
        company = (payload.get("company") or existing.company if existing else payload.get("company") or "").strip()
        location = (payload.get("location") or existing.location if existing else payload.get("location") or "").strip()
        job_type = (payload.get("job_type") or existing.job_type if existing else payload.get("job_type") or "").strip()

        if not title:
            raise ApiError("Job title is required.", 400)
        if not company:
            raise ApiError("Company name is required.", 400)
        if not location:
            raise ApiError("Job location is required.", 400)
        if not job_type:
            raise ApiError("Job type is required.", 400)

        slug = (payload.get("slug") or (existing.slug if existing else "")).strip()
        if not slug:
            slug = slugify(f"{title} {company} {location}")

        deadline_value = payload.get("deadline")
        if deadline_value in (None, ""):
            deadline = existing.deadline if existing else None
        else:
            try:
                deadline = date.fromisoformat(str(deadline_value)[:10])
            except ValueError:
                raise ApiError("Invalid deadline. Use YYYY-MM-DD.", 400)

        return {
            "title": title,
            "slug": slug,
            "company": company,
            "logo_url": (payload.get("logo_url") or (existing.logo_url if existing else "") or "").strip() or None,
            "location": location,
            "salary": (payload.get("salary") or (existing.salary if existing else "") or "").strip() or None,
            "job_type": job_type,
            "deadline": deadline,
            "categories": _payload_list(payload, "categories") or (existing.categories if existing else []),
            "summary": (payload.get("summary") or (existing.summary if existing else "") or "").strip() or None,
            "about_company": (payload.get("about_company") or (existing.about_company if existing else "") or "").strip() or None,
            "responsibilities": _payload_list(payload, "responsibilities") or (existing.responsibilities if existing else []),
            "requirements": _payload_list(payload, "requirements") or (existing.requirements if existing else []),
            "benefits": _payload_list(payload, "benefits") or (existing.benefits if existing else []),
            "application_email": (payload.get("application_email") or (existing.application_email if existing else "") or "").strip() or None,
            "email_subject": (payload.get("email_subject") or (existing.email_subject if existing else "") or "").strip() or None,
            "how_to_apply": (payload.get("how_to_apply") or (existing.how_to_apply if existing else "") or "").strip() or None,
            "featured": _payload_bool(payload, "featured", existing.featured if existing else False),
            "active": _payload_bool(payload, "active", existing.active if existing else True),
            "immediate_start": _payload_bool(payload, "immediate_start", existing.immediate_start if existing else False),
        }

    DOCUMENT_LABELS = {
        "sales_order": ("Sales Order", "SO"),
        "invoice": ("Invoice", "INV"),
        "payment_receipt": ("Payment Receipt", "RCT"),
    }

    def order_document_number(order: Order, document_type: str) -> str:
        if document_type not in DOCUMENT_LABELS:
            raise ApiError("Unsupported document type.", 400)
        field_name = {
            "sales_order": "sales_order_number",
            "invoice": "invoice_number",
            "payment_receipt": "receipt_number",
        }[document_type]
        existing = getattr(order, field_name, None)
        if existing:
            return existing
        prefix = DOCUMENT_LABELS[document_type][1]
        number = f"{prefix}-{order.order_number}"
        setattr(order, field_name, number)
        return number

    def ensure_order_documents(order: Order) -> None:
        required_types = ["sales_order", "invoice"]
        if order.payment_status == "confirmed":
            required_types.append("payment_receipt")

        for document_type in required_types:
            number = order_document_number(order, document_type)
            existing = (
                db_session()
                .query(DocumentRecord)
                .filter(
                    DocumentRecord.order_id == order.id,
                    DocumentRecord.document_type == document_type,
                )
                .first()
            )
            if existing:
                existing.document_number = existing.document_number or number
                continue

            db_session().add(
                DocumentRecord(
                    order_id=order.id,
                    document_type=document_type,
                    document_number=number,
                    status="generated",
                )
            )
        db_session().flush()

    def record_inventory_movement(
        *,
        product: Product,
        movement_type: str,
        quantity: int,
        previous_stock: int,
        new_stock: int,
        order: Order | None = None,
        actor: User | None = None,
        note: str | None = None,
    ) -> None:
        db_session().add(
            InventoryMovement(
                product_id=product.id,
                order_id=order.id if order else None,
                actor_user_id=actor.id if actor else None,
                movement_type=movement_type,
                quantity=quantity,
                previous_stock=previous_stock,
                new_stock=new_stock,
                note=note,
            )
        )

    def _pdf_text(value) -> str:
        return (
            str(value if value is not None else "")
            .replace("\\", "\\\\")
            .replace("(", "\\(")
            .replace(")", "\\)")
        )

    def _wrap_pdf_line(line: str, width: int = 92) -> list[str]:
        words = str(line).split()
        if not words:
            return [""]
        lines: list[str] = []
        current = words[0]
        for word in words[1:]:
            if len(current) + len(word) + 1 > width:
                lines.append(current)
                current = word
            else:
                current = f"{current} {word}"
        lines.append(current)
        return lines

    def build_simple_pdf(title: str, lines: list[str]) -> bytes:
        rendered_lines: list[str] = []
        for line in lines:
            rendered_lines.extend(_wrap_pdf_line(line))

        text_ops = [
            "BT",
            "/F1 18 Tf",
            "50 760 Td",
            f"({_pdf_text(title)}) Tj",
            "/F1 10 Tf",
        ]
        line_count = 0
        for line in rendered_lines:
            if line_count >= 42:
                text_ops.append("0 -16 Td")
                text_ops.append("(Report truncated. Export Excel for the complete data.) Tj")
                break
            text_ops.append("0 -16 Td")
            text_ops.append(f"({_pdf_text(line)}) Tj")
            line_count += 1
        text_ops.append("ET")
        stream = "\n".join(text_ops).encode("latin-1", errors="replace")

        objects = [
            b"<< /Type /Catalog /Pages 2 0 R >>",
            b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
            b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
            b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
            b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream",
        ]

        pdf = bytearray(b"%PDF-1.4\n")
        offsets = [0]
        for index, obj in enumerate(objects, start=1):
            offsets.append(len(pdf))
            pdf.extend(f"{index} 0 obj\n".encode("ascii"))
            pdf.extend(obj)
            pdf.extend(b"\nendobj\n")
        xref_at = len(pdf)
        pdf.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
        pdf.extend(b"0000000000 65535 f \n")
        for offset in offsets[1:]:
            pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
        pdf.extend(
            f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_at}\n%%EOF\n".encode("ascii")
        )
        return bytes(pdf)

    def order_document_lines(order: Order, document_type: str) -> list[str]:
        if document_type == "payment_receipt" and order.payment_status != "confirmed":
            raise ApiError("Payment receipt is only available after payment confirmation.", 400)

        ensure_order_documents(order)
        shipping = order.shipping_address or {}
        payment_details = order.payment_details or {}
        label = DOCUMENT_LABELS[document_type][0]
        number = order_document_number(order, document_type)
        lines = [
            "Hoinam Energy",
            "235 Umuocham Road, off Tonimas Junction by Enugu-PHC Express, Osisioma, Aba, Abia.",
            f"{label} number: {number}",
            f"Order number: {order.order_number}",
            f"Date: {order.created_at.strftime('%Y-%m-%d')}",
            f"Customer: {shipping.get('full_name') or order.user.full_name if order.user else 'Customer'}",
            f"Phone: {shipping.get('phone') or ''}",
            f"Address: {', '.join(str(shipping.get(key) or '') for key in ['address', 'city', 'state']).strip(', ')}",
            f"Payment method: {payment_details.get('label') or order.payment_method}",
            f"Payment status: {order.payment_status}",
            f"Payment reference: {order.payment_reference}",
            "",
            "Items:",
        ]
        for item in order.items or []:
            lines.append(
                f"- {item.get('name', 'Product')} | Qty {item.get('quantity', 0)} | Unit {item.get('currency', order.currency)} {float(item.get('unit_price') or 0):,.2f} | Line {item.get('currency', order.currency)} {float(item.get('line_total') or 0):,.2f}"
            )
        lines.extend(
            [
                "",
                f"Total: {order.currency} {float(order.total_amount):,.2f}",
                "Generated by Hoinam Energy admin system.",
            ]
        )
        return lines

    def pdf_response(filename: str, title: str, lines: list[str]):
        return Response(
            build_simple_pdf(title, lines),
            mimetype="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    def build_admin_report_payload() -> dict:
        month_start = datetime.now(timezone.utc).replace(
            day=1, hour=0, minute=0, second=0, microsecond=0
        )

        def as_utc(value: datetime | None) -> datetime | None:
            if value is None:
                return None
            if value.tzinfo is None:
                return value.replace(tzinfo=timezone.utc)
            return value.astimezone(timezone.utc)

        orders = (
            db_session()
            .query(Order)
            .options(joinedload(Order.user))
            .order_by(Order.created_at.desc())
            .all()
        )
        products = (
            db_session()
            .query(Product)
            .filter(Product.active.is_(True))
            .order_by(Product.name.asc())
            .all()
        )
        feedback_items = (
            db_session()
            .query(Feedback)
            .order_by(Feedback.created_at.desc())
            .all()
        )
        movements = (
            db_session()
            .query(InventoryMovement)
            .options(joinedload(InventoryMovement.product), joinedload(InventoryMovement.order))
            .order_by(InventoryMovement.created_at.desc())
            .limit(250)
            .all()
        )

        confirmed_orders = [order for order in orders if order.payment_status == "confirmed"]
        total_revenue = sum((order.total_amount for order in confirmed_orders), Decimal("0.00"))
        monthly_revenue = sum(
            (
                order.total_amount
                for order in confirmed_orders
                if as_utc(order.created_at) and as_utc(order.created_at) >= month_start
            ),
            Decimal("0.00"),
        )
        outstanding_orders = [
            order for order in orders if order.payment_status != "confirmed"
        ]

        product_performance: dict[str, dict] = {}
        for order in orders:
            for item in order.items or []:
                name = item.get("name") or f"Product {item.get('product_id')}"
                record = product_performance.setdefault(
                    name,
                    {
                        "product_name": name,
                        "quantity_sold": 0,
                        "gross_revenue": 0.0,
                        "confirmed_revenue": 0.0,
                        "orders": 0,
                    },
                )
                quantity = int(item.get("quantity") or 0)
                line_total = float(item.get("line_total") or 0)
                record["quantity_sold"] += quantity
                record["gross_revenue"] += line_total
                record["orders"] += 1
                if order.payment_status == "confirmed":
                    record["confirmed_revenue"] += line_total

        low_stock = [
            product
            for product in products
            if 0 <= int(product.stock or 0) <= 3
        ]

        return {
            "summary": {
                "total_sales": len(orders),
                "confirmed_sales": len(confirmed_orders),
                "total_revenue": float(total_revenue),
                "monthly_revenue": float(monthly_revenue),
                "outstanding_invoices": len(outstanding_orders),
                "customer_enquiries": len(feedback_items),
                "new_customer_enquiries": len([item for item in feedback_items if item.status == "new"]),
                "currency": settings.default_currency,
            },
            "inventory_status": {
                "active_products": len(products),
                "total_units": sum(max(0, int(product.stock or 0)) for product in products),
                "low_stock": len([product for product in products if 0 < int(product.stock or 0) <= 3]),
                "out_of_stock": len([product for product in products if int(product.stock or 0) == 0]),
            },
            "low_stock": [product.to_dict() for product in low_stock],
            "inventory_products": [product.to_dict() for product in products],
            "sales_history": [
                {
                    **order.to_dict(),
                    "user": order.user.to_dict() if order.user else None,
                }
                for order in orders[:100]
            ],
            "outstanding_invoices": [
                {
                    **order.to_dict(),
                    "user": order.user.to_dict() if order.user else None,
                }
                for order in outstanding_orders[:100]
            ],
            "customer_enquiries": [item.to_dict() for item in feedback_items[:100]],
            "inventory_history": [movement.to_dict() for movement in movements],
            "stock_movements": [movement.to_dict() for movement in movements],
            "product_performance": sorted(
                product_performance.values(),
                key=lambda item: (item["quantity_sold"], item["gross_revenue"]),
                reverse=True,
            )[:50],
        }

    def admin_report_pdf(report: dict) -> bytes:
        summary = report["summary"]
        inventory = report["inventory_status"]
        lines = [
            f"Total sales: {summary['total_sales']}",
            f"Confirmed sales: {summary['confirmed_sales']}",
            f"Total revenue: {summary['currency']} {summary['total_revenue']:,.2f}",
            f"Monthly revenue: {summary['currency']} {summary['monthly_revenue']:,.2f}",
            f"Outstanding invoices: {summary['outstanding_invoices']}",
            f"Customer enquiries: {summary['customer_enquiries']} ({summary['new_customer_enquiries']} new)",
            "",
            f"Inventory: {inventory['active_products']} active products, {inventory['total_units']} units",
            f"Low stock: {inventory['low_stock']} | Out of stock: {inventory['out_of_stock']}",
            "",
            "Top products:",
        ]
        for item in report["product_performance"][:10]:
            lines.append(
                f"- {item['product_name']}: {item['quantity_sold']} sold, gross {summary['currency']} {item['gross_revenue']:,.2f}"
            )
        lines.append("")
        lines.append("Low-stock alerts:")
        for product in report["low_stock"][:12]:
            lines.append(f"- {product['name']}: {product['stock']} in stock")
        return build_simple_pdf("Hoinam Energy Admin Report", lines)

    @app.get("/api/health")
    def health_check():
        return json_success({"status": "ok"})

    @app.get("/api/health/database")
    def health_database():
        """Check database health and data integrity."""
        try:
            session = db_session()
            
            # Count products and check for data issues
            total_products = session.query(func.count(Product.id)).scalar() or 0
            
            # Count products with missing critical fields
            products_missing_name = session.query(func.count(Product.id)).filter(
                Product.name == None
            ).scalar() or 0
            products_missing_slug = session.query(func.count(Product.id)).filter(
                Product.slug == None
            ).scalar() or 0
            products_missing_image = session.query(func.count(Product.id)).filter(
                Product.image_url == None
            ).scalar() or 0
            products_with_negative_stock = session.query(func.count(Product.id)).filter(
                Product.stock < 0
            ).scalar() or 0
            
            # Calculate integrity score
            issues = (
                products_missing_name + 
                products_missing_slug + 
                products_with_negative_stock
            )
            integrity_score = 100 if issues == 0 else max(0, 100 - (issues * 10))
            
            return json_success(
                {
                    "status": "ok",
                    "database": {
                        "total_products": total_products,
                        "integrity_score": integrity_score,
                        "issues": {
                            "missing_name": products_missing_name,
                            "missing_slug": products_missing_slug,
                            "missing_image": products_missing_image,
                            "negative_stock": products_with_negative_stock,
                        },
                    },
                }
            )
        except Exception as e:
            return json_success(
                {
                    "status": "error",
                    "error": str(e),
                    "database": {
                        "total_products": 0,
                        "integrity_score": 0,
                    },
                },
                status_code=500,
            )

    @app.get("/api/debug/smtp-config")
    def debug_smtp_config():
        """Check SMTP configuration (admin only)."""
        user = authenticate()
        if user.role != "admin":
            raise ApiError("Admin access required.", 403)

        return json_success(
            {
                "smtp_configured": smtp_is_configured(settings),
                "smtp_host": settings.smtp_host,
                "smtp_port": settings.smtp_port,
                "smtp_username": (
                    settings.smtp_username[:10] + "***"
                    if settings.smtp_username
                    else None
                ),
                "smtp_password": "***" if settings.smtp_password else None,
                "smtp_from_email": settings.smtp_from_email,
                "smtp_use_tls": settings.smtp_use_tls,
                "smtp_timeout_seconds": settings.smtp_timeout_seconds,
                "order_notification_email": settings.order_notification_email,
            }
        )

    @app.post("/api/debug/send-test-email")
    def send_test_email():
        """Send a test email (admin only). Pass {"to": "email@example.com"} to override recipient."""
        user = authenticate()
        if user.role != "admin":
            raise ApiError("Admin access required.", 403)

        if not smtp_is_configured(settings):
            raise ApiError(
                "SMTP is not configured. Check your environment variables.", 400
            )

        payload = request.get_json(silent=True) or {}
        recipient = (payload.get("to") or "").strip() or settings.order_notification_email

        try:
            test_message = EmailMessage()
            test_message["Subject"] = "[TEST] Hoinam Energy SMTP Test"
            test_message["From"] = settings.smtp_from_email or settings.smtp_username
            test_message["To"] = recipient
            test_message.set_content(
                f"SMTP test email from Hoinam Energy.\n\n"
                f"Sent at: {datetime.now(timezone.utc).isoformat()}\n"
                f"From: {settings.smtp_from_email or settings.smtp_username}\n"
                f"To: {recipient}\n\n"
                f"If you received this, SMTP is working correctly."
            )
            send_message_via_smtp(settings, test_message)
            return json_success(
                {"message": f"Test email sent to {recipient}"}
            )
        except Exception as e:
            app.logger.exception("Test email failed")
            raise ApiError(f"Failed to send test email: {str(e)}", 500)

    @app.post("/api/auth/verify")
    def verify_auth():
        user = authenticate()
        return json_success(user.to_dict())

    @app.put("/api/profile")
    def update_profile():
        user = authenticate()
        payload = request.get_json(silent=True) or {}

        if "phone" in payload:
            phone = (payload.get("phone") or "").strip()
            if not phone:
                raise ApiError("Phone number is required.", 400)
            user.phone = phone

        if "full_name" in payload:
            full_name = (payload.get("full_name") or "").strip()
            if full_name:
                user.full_name = full_name

        flag_duplicate_full_name_users(db_session(), user)
        db_session().commit()
        return json_success(user.to_dict())

    @app.get("/api/stores")
    def list_stores():
        stores = get_all_stores()
        return json_success([store.to_dict() for store in stores])

    @app.get("/api/stores/<store_slug>")
    def get_store(store_slug: str):
        store = get_store_by_slug(store_slug)
        if not store:
            raise ApiError("Store not found.", 404)
        return json_success(store.to_dict())

    @app.get("/api/products")
    def list_products():
        # Support filtering by store
        store_slug = request.args.get("store", "").strip()

        query = db_session().query(Product).filter(Product.active.is_(True))

        if store_slug:
            query = query.filter(Product.store_slug == store_slug)

        products = query.order_by(desc(Product.featured), Product.name.asc()).all()
        return json_success([product.to_dict() for product in products])

    @app.get("/api/products/<int:product_id>")
    def get_product(product_id: int):
        product = db_session().query(Product).filter(Product.id == product_id).first()
        if not product or not product.active:
            raise ApiError("Product not found.", 404)
        return json_success(product.to_dict())

    @app.post("/api/products")
    def create_product():
        user = authenticate(admin=True)
        payload = request.get_json(silent=True) or {}
        product = Product()
        apply_product_payload(product, payload)
        db_session().add(product)
        db_session().flush()
        record_inventory_movement(
            product=product,
            movement_type="adjustment",
            quantity=int(product.stock or 0),
            previous_stock=0,
            new_stock=int(product.stock or 0),
            actor=user,
            note="Product created from admin dashboard.",
        )
        db_session().commit()
        return json_success(product.to_dict(), status_code=201)

    @app.put("/api/products")
    @app.put("/api/products/<int:product_id>")
    def update_product(product_id: int | None = None):
        user = authenticate(admin=True)
        payload = request.get_json(silent=True) or {}
        target_id = product_id or payload.get("id")
        if not target_id:
            raise ApiError("Product id is required.", 400)
        product = db_session().query(Product).filter(Product.id == target_id).first()
        if not product:
            raise ApiError("Product not found.", 404)
        previous_stock = int(product.stock or 0)
        apply_product_payload(product, payload)
        new_stock = int(product.stock or 0)
        if previous_stock != new_stock:
            record_inventory_movement(
                product=product,
                movement_type="adjustment",
                quantity=new_stock - previous_stock,
                previous_stock=previous_stock,
                new_stock=new_stock,
                actor=user,
                note="Stock updated from admin product form.",
            )
        db_session().commit()
        return json_success(product.to_dict())

    @app.delete("/api/products/<int:product_id>")
    def delete_product(product_id: int):
        authenticate(admin=True)
        product = db_session().query(Product).filter(Product.id == product_id).first()
        if not product:
            raise ApiError("Product not found.", 404)
        product.active = False
        db_session().commit()
        return json_success({"id": product_id}, message="Product archived.")

    @app.get("/api/jobs")
    def list_jobs():
        jobs = (
            db_session()
            .query(JobListing)
            .filter(JobListing.active.is_(True))
            .order_by(desc(JobListing.featured), JobListing.created_at.desc())
            .all()
        )
        return json_success([job.to_dict() for job in jobs])

    @app.get("/api/jobs/<int:job_id>")
    def get_job(job_id: int):
        job = db_session().query(JobListing).filter(JobListing.id == job_id).first()
        if not job or not job.active:
            raise ApiError("Job listing not found.", 404)
        return json_success(job.to_dict())

    @app.get("/api/admin/jobs")
    def list_admin_jobs():
        authenticate(admin=True)
        jobs = (
            db_session()
            .query(JobListing)
            .order_by(desc(JobListing.featured), JobListing.created_at.desc())
            .all()
        )
        return json_success([job.to_dict() for job in jobs])

    @app.post("/api/admin/jobs")
    def create_job():
        authenticate(admin=True)
        payload = request.get_json(silent=True) or {}
        normalized = normalize_job_payload(payload)

        existing = (
            db_session()
            .query(JobListing)
            .filter(JobListing.slug == normalized["slug"])
            .first()
        )
        if existing:
            raise ApiError("A job with this slug already exists.", 409)

        job = JobListing(**normalized)
        db_session().add(job)
        db_session().commit()
        return json_success(job.to_dict(), status_code=201, message="Job listing created.")

    @app.put("/api/admin/jobs/<int:job_id>")
    def update_job(job_id: int):
        authenticate(admin=True)
        job = db_session().query(JobListing).filter(JobListing.id == job_id).first()
        if not job:
            raise ApiError("Job listing not found.", 404)

        payload = request.get_json(silent=True) or {}
        normalized = normalize_job_payload(payload, existing=job)
        duplicate = (
            db_session()
            .query(JobListing)
            .filter(JobListing.slug == normalized["slug"], JobListing.id != job_id)
            .first()
        )
        if duplicate:
            raise ApiError("Another job with this slug already exists.", 409)

        for field_name, value in normalized.items():
            setattr(job, field_name, value)
        db_session().commit()
        return json_success(job.to_dict(), message="Job listing updated.")

    @app.delete("/api/admin/jobs/<int:job_id>")
    def delete_job(job_id: int):
        authenticate(admin=True)
        job = db_session().query(JobListing).filter(JobListing.id == job_id).first()
        if not job:
            raise ApiError("Job listing not found.", 404)

        job.active = False
        db_session().commit()
        return json_success({"id": job_id}, message="Job listing deactivated.")

    def payment_options_payload() -> dict:
        return {
            "methods": [
                {
                    "id": "bank_transfer",
                    "kind": "transfer",
                    "label": "Transfer before delivery",
                    "description": "Place the order now, then pay by direct bank transfer to the Hoinam Energy company account. You'll receive a verification code to include in the transfer narration.",
                    "bank_name": settings.bank_transfer_bank_name,
                    "account_number": settings.bank_transfer_account_number,
                    "account_name": settings.bank_transfer_account_name,
                },
                {
                    "id": "pay_on_delivery",
                    "kind": "delivery",
                    "label": "Pay on delivery",
                    "description": "Reserve the order and pay when Hoinam Energy confirms delivery.",
                },
            ],
            "unavailable": [
                {
                    "id": "opay_transfer",
                    "label": "OPay merchant",
                    "reason": "Under maintenance — coming soon.",
                },
            ],
        }

    def payment_option_by_id(payment_method: str) -> dict | None:
        methods = payment_options_payload().get("methods") or []
        for option in methods:
            if option.get("id") == payment_method:
                return option
        return None

    def payment_details_payload(payment_method: str) -> dict:
        option = payment_option_by_id(payment_method)
        if option is None:
            return {
                "id": payment_method,
                "kind": (
                    "transfer" if payment_method != "pay_on_delivery" else "delivery"
                ),
                "label": payment_method.replace("_", " ").title(),
            }

        payload = {
            "id": option["id"],
            "kind": option.get("kind")
            or ("transfer" if option["id"] != "pay_on_delivery" else "delivery"),
            "label": option.get("label"),
            "description": option.get("description"),
        }

        if payload["kind"] == "transfer":
            payload.update(
                {
                    "bank_name": option.get("bank_name"),
                    "account_number": option.get("account_number"),
                    "account_name": option.get("account_name"),
                }
            )

        return payload

    @app.get("/api/payment-options")
    def payment_options_route():
        return json_success(payment_options_payload())

    @app.post("/api/orders")
    def create_order():
        user = authenticate()
        payload = request.get_json(silent=True) or {}
        payment_method = (payload.get("payment_method") or "bank_transfer").strip()
        allowed_payment_methods = {"bank_transfer", "pay_on_delivery"}
        if payment_method not in allowed_payment_methods:
            raise ApiError(
                "Choose transfer before delivery or pay on delivery.", 400
            )

        payment_reference = (
            payload.get("payment_reference") or ""
        ).strip() or generate_payment_reference(
            "BANK" if payment_method == "bank_transfer" else "POD"
        )
        existing_order = (
            db_session()
            .query(Order)
            .filter(Order.payment_reference == payment_reference)
            .first()
        )
        if existing_order:
            ensure_order_documents(existing_order)
            db_session().commit()
            return json_success(existing_order.to_dict())

        shipping_address = payload.get("shipping_address") or {}
        required_fields = ["full_name", "phone", "address", "city", "state"]
        missing_fields = [
            field for field in required_fields if not shipping_address.get(field)
        ]
        if missing_fields:
            raise ApiError(
                f"Missing shipping fields: {', '.join(missing_fields)}.", 400
            )

        try:
            normalized_items, total, locked_products = calculate_order_items(
                db_session(), payload.get("items") or [], lock_products=True
            )
        except StockError as stock_err:
            # Return a structured stock error so the frontend can show the modal
            return jsonify({
                "success": False,
                "error_type": "stock_error",
                "message": f"Not enough stock for {stock_err.product_name}.",
                "product_name": stock_err.product_name,
                "requested": stock_err.requested,
                "available": stock_err.available,
                "discount_code": "SORRY2",
            }), 409

        stock_movements = []
        for item in normalized_items:
            product = locked_products[item["product_id"]]
            previous_stock = int(product.stock or 0)
            product.stock = max(0, previous_stock - int(item["quantity"]))
            stock_movements.append((product, previous_stock, int(product.stock or 0), int(item["quantity"])))

        payment_status = (
            "awaiting_transfer"
            if payment_method == "bank_transfer"
            else "pay_on_delivery"
        )
        order_status = (
            "payment_pending"
            if payment_method == "bank_transfer"
            else "confirmed"
        )

        # ── Coupon validation ─────────────────────────────────────────────────
        coupon_code = (payload.get("coupon_code") or "").strip().upper()
        coupon = None
        discount_amount = Decimal("0.00")
        coupon_error = None

        if coupon_code:
            coupon = (
                db_session()
                .query(Coupon)
                .filter(
                    Coupon.code == coupon_code,
                    Coupon.is_active.is_(True),
                )
                .first()
            )
            if not coupon:
                coupon_error = "Coupon code not found or is no longer active."
            elif coupon.expires_at and coupon.expires_at < datetime.now(timezone.utc):
                coupon_error = "This coupon has expired."
            elif coupon.max_uses is not None and coupon.uses >= coupon.max_uses:
                coupon_error = "This coupon has reached its usage limit."
            elif total < coupon.min_order_amount:
                coupon_error = (
                    f"This coupon requires a minimum order of "
                    f"{float(coupon.min_order_amount):,.2f} {settings.default_currency}."
                )
            else:
                if coupon.discount_type == "percent":
                    discount_amount = (total * coupon.discount_value / Decimal("100")).quantize(
                        Decimal("0.01")
                    )
                else:
                    discount_amount = min(coupon.discount_value, total)

            if coupon_error:
                return jsonify({
                    "success": False,
                    "error_type": "coupon_error",
                    "message": coupon_error,
                }), 400

        final_total = max(total - discount_amount, Decimal("0.00"))

        order = Order(
            order_number=generate_order_number(),
            user_id=user.id,
            status=order_status,
            payment_status=payment_status,
            payment_method=payment_method,
            payment_reference=payment_reference,
            payment_details={
                **payment_details_payload(payment_method),
                **({"coupon_code": coupon_code, "discount_amount": float(discount_amount)} if coupon else {}),
            },
            total_amount=to_decimal(final_total),
            currency=settings.default_currency,
            shipping_address=shipping_address,
            notes=payload.get("notes"),
            items=normalized_items,
            sales_channel="website",
        )
        db_session().add(order)
        db_session().flush()

        ensure_order_documents(order)
        for product, previous_stock, new_stock, quantity in stock_movements:
            record_inventory_movement(
                product=product,
                movement_type="sale",
                quantity=-quantity,
                previous_stock=previous_stock,
                new_stock=new_stock,
                order=order,
                actor=user,
                note=f"Website order {order.order_number}.",
            )

        # Increment coupon usage
        if coupon:
            coupon.uses += 1

        # Generate verification code for bank transfers
        verification_code = None
        if payment_method == "bank_transfer":
            verification_code = generate_verification_code()
            payment = Payment(
                order_id=order.id,
                verification_code=verification_code,
                amount=to_decimal(final_total),
                currency=settings.default_currency,
                payment_method=payment_method,
                status="pending",
            )
            db_session().add(payment)
            # Store verification code in payment_details so dashboard can access it
            details = dict(order.payment_details or {})
            details["verification_code"] = verification_code
            order.payment_details = details

        db_session().commit()

        if smtp_is_configured(settings):
            try:
                message = build_order_notification_message(
                    settings=settings,
                    user=user,
                    order=order,
                    shipping_address=shipping_address,
                    verification_code=verification_code,
                )
                send_message_via_smtp(settings, message)
            except Exception:
                app.logger.exception(
                    "Order %s was created but the notification email could not be sent.",
                    order.order_number,
                )
        else:
            app.logger.warning(
                "SMTP is not configured, so no order notification email was sent for %s.",
                order.order_number,
            )

        order_payload = order.to_dict()
        order_payload["payment_options"] = payment_options_payload()
        if verification_code:
            order_payload["verification_code"] = verification_code
        return json_success(order_payload, status_code=201)

    @app.post("/api/admin/sales-orders")
    def create_admin_sales_order():
        admin_user = authenticate(admin=True)
        payload = request.get_json(silent=True) or {}

        raw_items = payload.get("items") or []
        if not raw_items and payload.get("product_id"):
            raw_items = [
                {
                    "product_id": payload.get("product_id"),
                    "quantity": payload.get("quantity") or 1,
                }
            ]

        customer_payload = payload.get("customer") or {}
        full_name = (
            customer_payload.get("full_name")
            or payload.get("customer_name")
            or ""
        ).strip()
        if not full_name:
            raise ApiError("Customer name is required.", 400)

        email = (
            customer_payload.get("email")
            or payload.get("customer_email")
            or ""
        ).strip().lower() or None
        phone = (
            customer_payload.get("phone")
            or payload.get("customer_phone")
            or ""
        ).strip() or None

        customer = None
        if payload.get("user_id"):
            customer = (
                db_session()
                .query(User)
                .filter(User.id == int(payload["user_id"]))
                .first()
            )
        if customer is None and email:
            customer = (
                db_session()
                .query(User)
                .filter(User.email == email)
                .first()
            )
        if customer is None:
            customer = User(
                firebase_uid=generate_payment_reference("ADMINCUSTOMER"),
                email=email,
                full_name=full_name,
                phone=phone,
                role="user",
            )
            db_session().add(customer)
            db_session().flush()
        else:
            customer.full_name = full_name or customer.full_name
            customer.phone = phone or customer.phone

        shipping_address = {
            "full_name": full_name,
            "phone": phone or customer.phone or "",
            "address": (
                customer_payload.get("address")
                or payload.get("address")
                or "Admin-created sale"
            ),
            "city": customer_payload.get("city") or payload.get("city") or "Aba",
            "state": customer_payload.get("state") or payload.get("state") or "Abia",
        }

        payment_method = (payload.get("payment_method") or "bank_transfer").strip()
        if payment_method not in {"bank_transfer", "pay_on_delivery"}:
            raise ApiError("Choose bank transfer or pay on delivery.", 400)

        try:
            normalized_items, total, locked_products = calculate_order_items(
                db_session(), raw_items, lock_products=True
            )
        except StockError as stock_err:
            return jsonify({
                "success": False,
                "error_type": "stock_error",
                "message": f"Not enough stock for {stock_err.product_name}.",
                "product_name": stock_err.product_name,
                "requested": stock_err.requested,
                "available": stock_err.available,
            }), 409
        except ValueError as exc:
            raise ApiError(str(exc), 400) from exc

        stock_movements = []
        for item in normalized_items:
            product = locked_products[item["product_id"]]
            previous_stock = int(product.stock or 0)
            product.stock = max(0, previous_stock - int(item["quantity"]))
            stock_movements.append((product, previous_stock, int(product.stock or 0), int(item["quantity"])))

        payment_confirmed = _payload_bool(payload, "payment_confirmed", False)
        payment_status = (
            "confirmed"
            if payment_confirmed
            else ("awaiting_transfer" if payment_method == "bank_transfer" else "pay_on_delivery")
        )
        order_status = (
            "delivered"
            if payment_confirmed
            else ("payment_pending" if payment_method == "bank_transfer" else "confirmed")
        )
        payment_reference = (
            payload.get("payment_reference") or ""
        ).strip() or generate_payment_reference("ADMIN")

        order = Order(
            order_number=generate_order_number(),
            user_id=customer.id,
            status=order_status,
            payment_status=payment_status,
            payment_method=payment_method,
            payment_reference=payment_reference,
            payment_details=payment_details_payload(payment_method),
            total_amount=to_decimal(total),
            currency=settings.default_currency,
            shipping_address=shipping_address,
            notes=(payload.get("notes") or "").strip() or "Created from admin sales order form.",
            items=normalized_items,
            sales_channel="admin",
            created_by_id=admin_user.id,
        )
        db_session().add(order)
        db_session().flush()
        ensure_order_documents(order)

        for product, previous_stock, new_stock, quantity in stock_movements:
            record_inventory_movement(
                product=product,
                movement_type="admin_sale",
                quantity=-quantity,
                previous_stock=previous_stock,
                new_stock=new_stock,
                order=order,
                actor=admin_user,
                note=f"Admin sales order {order.order_number}.",
            )

        if payment_method == "bank_transfer" or payment_confirmed:
            verification_code = generate_verification_code("ADM")
            payment = Payment(
                order_id=order.id,
                verification_code=verification_code,
                amount=to_decimal(total),
                currency=settings.default_currency,
                payment_method=payment_method,
                status="confirmed" if payment_confirmed else "pending",
                notes=(
                    f"Payment {'confirmed' if payment_confirmed else 'created'} by admin "
                    f"{admin_user.email or admin_user.id}."
                ),
            )
            db_session().add(payment)
            details = dict(order.payment_details or {})
            details["verification_code"] = verification_code
            order.payment_details = details

        db_session().commit()

        order_payload = order.to_dict()
        order_payload["user"] = customer.to_dict()
        order_payload["documents"] = [doc.to_dict() for doc in order.documents]
        return json_success(
            order_payload,
            status_code=201,
            message=f"Sales order {order.order_number} created.",
        )

    @app.get("/api/orders/user")
    def get_user_orders():
        user = authenticate()
        orders = (
            db_session()
            .query(Order)
            .filter(Order.user_id == user.id)
            .order_by(Order.created_at.desc())
            .all()
        )
        return json_success([order.to_dict() for order in orders])

    @app.get("/api/orders")
    def get_all_orders():
        authenticate(admin=True)
        orders = (
            db_session()
            .query(Order)
            .options(joinedload(Order.user))
            .order_by(Order.created_at.desc())
            .all()
        )
        payload = []
        for order in orders:
            data = order.to_dict()
            data["user"] = order.user.to_dict() if order.user else None
            payload.append(data)
        return json_success(payload)

    @app.get("/api/orders/<int:order_id>/documents/<document_type>")
    def download_order_document(order_id: int, document_type: str):
        user = authenticate()
        if document_type not in DOCUMENT_LABELS:
            raise ApiError("Unsupported document type.", 400)

        order = (
            db_session()
            .query(Order)
            .options(joinedload(Order.user))
            .filter(Order.id == order_id)
            .first()
        )
        if not order:
            raise ApiError("Order not found.", 404)
        if user.role != "admin" and order.user_id != user.id:
            raise ApiError("You do not have access to this order document.", 403)

        label = DOCUMENT_LABELS[document_type][0]
        number = order_document_number(order, document_type)
        lines = order_document_lines(order, document_type)
        db_session().commit()
        safe_number = number.lower().replace("/", "-")
        return pdf_response(
            f"{safe_number}.pdf",
            f"Hoinam Energy {label}",
            lines,
        )

    @app.post("/api/payments/<verification_code>/receipt")
    def upload_payment_receipt(verification_code: str):
        user = authenticate()

        # Find the payment by verification code
        payment = (
            db_session()
            .query(Payment)
            .filter(Payment.verification_code == verification_code)
            .first()
        )
        if not payment:
            raise ApiError("Payment verification code not found.", 404)

        # Verify the order belongs to the authenticated user
        order = payment.order
        if order.user_id != user.id:
            raise ApiError(
                "You don't have permission to upload receipt for this payment.", 403
            )

        file_storage = request.files.get("receipt")
        if not file_storage:
            raise ApiError("Please attach a receipt file.", 400)

        # Handle different receipt formats (image or PDF)
        allowed_extensions = {".pdf", ".png", ".jpg", ".jpeg", ".gif"}
        file_extension = Path(file_storage.filename).suffix.lower()
        if file_extension not in allowed_extensions:
            raise ApiError("Allowed formats: PDF, PNG, JPG, JPEG, GIF", 400)

        # Store receipt URL or file path
        receipt_filename = f"receipt_{verification_code}_{file_storage.filename}"
        receipt_path = project_root / "receipts" / receipt_filename
        receipt_path.parent.mkdir(parents=True, exist_ok=True)
        file_storage.save(receipt_path)

        from datetime import datetime as dt

        payment.receipt_url = f"/receipts/{receipt_filename}"
        payment.receipt_uploaded_at = dt.now(timezone.utc)
        payment.status = "receipt_uploaded"
        order.payment_status = "receipt_uploaded"
        order.status = "payment_pending"
        db_session().commit()

        # Notify admin that a receipt was uploaded and needs review
        if smtp_is_configured(settings):
            try:
                from email.message import EmailMessage as _EM
                admin_msg = _EM()
                admin_msg["Subject"] = f"Receipt uploaded for order {order.order_number} — needs review"
                admin_msg["From"] = settings.smtp_from_email or settings.smtp_username
                admin_msg["To"] = settings.order_notification_email
                admin_msg.set_content(
                    f"A customer has uploaded a payment receipt for order {order.order_number}.\n\n"
                    f"Order ID: {order.id}\n"
                    f"Order number: {order.order_number}\n"
                    f"Payment reference: {order.payment_reference}\n"
                    f"Total: {float(order.total_amount):,.2f} {order.currency}\n"
                    f"Verification code: {payment.verification_code}\n\n"
                    f"Please review and approve the order in the admin panel."
                )
                send_message_via_smtp(settings, admin_msg)
            except Exception:
                app.logger.exception(
                    "Receipt uploaded for %s but admin notification email failed.",
                    order.order_number,
                )

        return json_success(payment.to_dict(), message="Receipt uploaded. Hoinam Energy will verify and approve your order within 40 minutes.")

    @app.post("/api/installations")
    def create_installation():
        user = authenticate()
        payload = request.get_json(silent=True) or {}
        address = (payload.get("address") or "").strip()
        if not address:
            raise ApiError("Installation address is required.", 400)

        preferred_date = payload.get("preferred_date")
        installation = Installation(
            user_id=user.id,
            product_id=payload.get("product_id"),
            preferred_date=(
                date.fromisoformat(preferred_date) if preferred_date else None
            ),
            service_type=payload.get("service_type") or "Solar Installation",
            address=address,
            city=payload.get("city"),
            state=payload.get("state"),
            phone=payload.get("phone") or user.phone,
            notes=payload.get("notes"),
            status="pending",
        )
        db_session().add(installation)
        db_session().commit()
        return json_success(installation.to_dict(), status_code=201)

    @app.get("/api/installations/user")
    def get_user_installations():
        user = authenticate()
        installations = (
            db_session()
            .query(Installation)
            .filter(Installation.user_id == user.id)
            .order_by(Installation.created_at.desc())
            .all()
        )
        return json_success([serialize_installation(item) for item in installations])

    @app.get("/api/installations")
    def get_all_installations():
        authenticate(admin=True)
        installations = (
            db_session()
            .query(Installation)
            .options(joinedload(Installation.user), joinedload(Installation.product))
            .order_by(Installation.created_at.desc())
            .all()
        )
        return json_success([serialize_installation(item) for item in installations])

    @app.put("/api/installations/<int:installation_id>")
    def update_installation(installation_id: int):
        authenticate(admin=True)
        installation = (
            db_session()
            .query(Installation)
            .filter(Installation.id == installation_id)
            .first()
        )
        if not installation:
            raise ApiError("Installation request not found.", 404)

        payload = request.get_json(silent=True) or {}
        if "status" in payload:
            installation.status = payload["status"]
        if "assigned_to" in payload:
            installation.assigned_to = payload["assigned_to"]
        if "preferred_date" in payload and payload["preferred_date"]:
            installation.preferred_date = date.fromisoformat(payload["preferred_date"])
        db_session().commit()
        return json_success(serialize_installation(installation))

    @app.get("/api/users")
    def list_users():
        authenticate(admin=True)
        users = db_session().query(User).order_by(User.created_at.desc()).all()
        return json_success([user.to_dict() for user in users])

    def _notify_subscribers_of_new_post(post: BlogPost) -> None:
        """Email all active subscribers about a newly published blog post."""
        if not smtp_is_configured(settings):
            app.logger.warning("SMTP not configured — skipping subscriber notifications for %s.", post.slug)
            return

        subscribers = (
            db_session()
            .query(BlogSubscriber)
            .filter(BlogSubscriber.is_active.is_(True))
            .all()
        )
        if not subscribers:
            return

        post_url = f"{settings.frontend_url}/blog-post.html?slug={post.slug}"
        sent = 0
        failed = 0
        for subscriber in subscribers:
            try:
                msg = build_new_post_notification_message(
                    settings=settings,
                    subscriber_email=subscriber.email,
                    subscriber_name=subscriber.name,
                    post_title=post.title,
                    post_excerpt=post.excerpt or "",
                    post_url=post_url,
                    unsubscribe_token=subscriber.unsubscribe_token,
                    frontend_url=settings.frontend_url,
                )
                send_message_via_smtp(settings, msg)
                sent += 1
            except Exception:
                app.logger.exception(
                    "Failed to send new-post notification to %s for post %s.",
                    subscriber.email,
                    post.slug,
                )
                failed += 1

        app.logger.info(
            "New-post notifications for '%s': %d sent, %d failed.",
            post.title,
            sent,
            failed,
        )

    @app.get("/api/blog")
    def list_blog_posts():
        """Get all published blog posts (public)."""
        posts = (
            db_session()
            .query(BlogPost)
            .filter(BlogPost.is_published.is_(True))
            .order_by(BlogPost.published_at.desc())
            .all()
        )
        return json_success([post.to_dict() for post in posts])

    @app.get("/api/blog/unsubscribe/<token>")
    def blog_unsubscribe(token: str):
        """Unsubscribe via token link.

        NOTE: This route MUST be registered before GET /api/blog/<post_slug>
        so Flask doesn't swallow it as a wildcard slug match.
        """
        subscriber = (
            db_session()
            .query(BlogSubscriber)
            .filter(BlogSubscriber.unsubscribe_token == token)
            .first()
        )
        if not subscriber:
            raise ApiError("Unsubscribe link not found or already used.", 404)
        subscriber.is_active = False
        db_session().commit()
        return json_success({"email": subscriber.email}, message="You've been unsubscribed.")

    @app.get("/api/blog/<post_slug>")
    def get_blog_post(post_slug: str):
        """Get a single blog post by slug.
        Published posts are public. Draft posts are visible to admins only.
        """
        post = (
            db_session()
            .query(BlogPost)
            .filter(BlogPost.slug == post_slug)
            .first()
        )
        if not post:
            raise ApiError("Blog post not found.", 404)

        if not post.is_published:
            # Allow admins to preview drafts; everyone else gets a clear message
            user = authenticate(required=False)
            if not user or user.role != "admin":
                raise ApiError(
                    "This post is not published yet. If you're the admin, log in to preview it.",
                    404,
                )

        return json_success(post.to_dict())

    @app.get("/api/admin/blog")
    def list_admin_blog_posts():
        """Get all blog posts for admin (published and drafts)."""
        authenticate(admin=True)
        posts = (
            db_session()
            .query(BlogPost)
            .order_by(BlogPost.created_at.desc())
            .all()
        )
        return json_success([post.to_dict() for post in posts])

    @app.post("/api/admin/blog")
    def create_blog_post():
        """Create a new blog post (admin only)."""
        user = authenticate(admin=True)
        payload = request.get_json(silent=True) or {}

        title = (payload.get("title") or "").strip()
        if not title:
            raise ApiError("Blog post title is required.", 400)

        slug = (payload.get("slug") or "").strip().strip("/") or slugify(title)
        excerpt = (payload.get("excerpt") or "").strip()
        content = (payload.get("content") or "").strip()
        if not content:
            raise ApiError("Blog post content is required.", 400)

        # Check for duplicate slug
        existing = db_session().query(BlogPost).filter(BlogPost.slug == slug).first()
        if existing:
            raise ApiError("A blog post with this slug already exists.", 409)

        post = BlogPost(
            title=title,
            slug=slug,
            excerpt=excerpt or title,
            content=content,
            image_url=payload.get("image_url"),
            author_id=user.id,
            is_published=payload.get("is_published", False),
            published_at=datetime.now(timezone.utc) if payload.get("is_published") else None,
            category=payload.get("category", "News"),
            tags=payload.get("tags", []),
        )
        db_session().add(post)
        db_session().commit()

        if post.is_published:
            try:
                _notify_subscribers_of_new_post(post)
            except Exception:
                app.logger.exception("Subscriber notification failed for new post %s.", post.slug)

        return json_success(post.to_dict(), status_code=201, message="Blog post created.")

    @app.put("/api/admin/blog/<int:post_id>")
    def update_blog_post(post_id: int):
        """Update a blog post (admin only)."""
        user = authenticate(admin=True)
        payload = request.get_json(silent=True) or {}

        post = db_session().query(BlogPost).filter(BlogPost.id == post_id).first()
        if not post:
            raise ApiError("Blog post not found.", 404)

        if "title" in payload:
            post.title = (payload.get("title") or "").strip() or post.title
        if "slug" in payload:
            slug = (payload.get("slug") or "").strip().strip("/")
            if slug and slug != post.slug:
                existing = (
                    db_session()
                    .query(BlogPost)
                    .filter(BlogPost.slug == slug, BlogPost.id != post_id)
                    .first()
                )
                if existing:
                    raise ApiError("A blog post with this slug already exists.", 409)
                post.slug = slug
        if "excerpt" in payload:
            post.excerpt = (payload.get("excerpt") or "").strip() or post.excerpt
        if "content" in payload:
            post.content = (payload.get("content") or "").strip() or post.content
        if "image_url" in payload:
            post.image_url = payload.get("image_url")
        if "category" in payload:
            post.category = payload.get("category", "News")
        if "tags" in payload:
            post.tags = payload.get("tags", [])

        # Handle publish/unpublish — track transition before mutating
        just_published = False
        if "is_published" in payload:
            was_published = post.is_published
            post.is_published = bool(payload.get("is_published"))
            if post.is_published and not was_published:
                post.published_at = datetime.now(timezone.utc)
                just_published = True
            elif not post.is_published:
                post.published_at = None

        db_session().commit()

        if just_published:
            try:
                _notify_subscribers_of_new_post(post)
            except Exception:
                app.logger.exception("Subscriber notification failed for post %s.", post.slug)

        return json_success(post.to_dict(), message="Blog post updated.")

    @app.delete("/api/admin/blog/<int:post_id>")
    def delete_blog_post(post_id: int):
        """Delete a blog post (admin only)."""
        authenticate(admin=True)
        post = db_session().query(BlogPost).filter(BlogPost.id == post_id).first()
        if not post:
            raise ApiError("Blog post not found.", 404)
        db_session().delete(post)
        db_session().commit()
        return json_success({"id": post_id}, message="Blog post deleted.")

    # ── Blog subscriptions ────────────────────────────────────────────────────

    @app.post("/api/blog/subscribe")
    def blog_subscribe():
        """Subscribe an email address to blog post notifications."""
        payload = request.get_json(silent=True) or {}
        email = (payload.get("email") or "").strip().lower()
        name = (payload.get("name") or "").strip() or None
        if not email or "@" not in email:
            raise ApiError("A valid email address is required.", 400)

        existing = (
            db_session()
            .query(BlogSubscriber)
            .filter(BlogSubscriber.email == email)
            .first()
        )
        if existing:
            if existing.is_active:
                return json_success(
                    existing.to_dict(),
                    message="You're already subscribed.",
                )
            # Re-activate a previously unsubscribed address
            existing.is_active = True
            existing.name = name or existing.name
            db_session().commit()
            subscriber = existing
        else:
            subscriber = BlogSubscriber(
                email=email,
                name=name,
                is_active=True,
                unsubscribe_token=generate_unsubscribe_token(),
            )
            db_session().add(subscriber)
            db_session().commit()

        if smtp_is_configured(settings):
            try:
                msg = build_subscription_confirmation_message(
                    settings=settings,
                    subscriber_email=subscriber.email,
                    subscriber_name=subscriber.name,
                    unsubscribe_token=subscriber.unsubscribe_token,
                    frontend_url=settings.frontend_url,
                )
                send_message_via_smtp(settings, msg)
            except Exception as email_err:
                app.logger.exception(
                    "Subscription confirmation email failed for %s.", email
                )
                # Return success for the subscription itself but flag the email issue
                return json_success(
                    subscriber.to_dict(),
                    status_code=201,
                    message=f"Subscribed! However, the confirmation email could not be sent ({email_err}). Your subscription is saved.",
                )
        else:
            app.logger.warning("SMTP not configured — no confirmation email sent for subscriber %s.", email)

        return json_success(
            subscriber.to_dict(),
            status_code=201,
            message="Subscribed! Check your inbox for a confirmation.",
        )

    @app.get("/api/admin/blog/subscribers")
    def list_blog_subscribers():
        """List all blog subscribers (admin only)."""
        authenticate(admin=True)
        subscribers = (
            db_session()
            .query(BlogSubscriber)
            .filter(BlogSubscriber.is_active.is_(True))
            .order_by(BlogSubscriber.created_at.desc())
            .all()
        )
        return json_success([s.to_dict() for s in subscribers])

    @app.get("/api/admin/stats")
    @app.get("/api/admin/stats")
    def admin_stats():
        authenticate(admin=True)

        # Revenue — confirmed orders only
        confirmed_revenue = db_session().query(
            func.coalesce(func.sum(Order.total_amount), 0)
        ).filter(Order.payment_status == "confirmed").scalar()

        # All-time revenue (including pending)
        total_revenue = db_session().query(
            func.coalesce(func.sum(Order.total_amount), 0)
        ).scalar()

        # Order counts by status
        orders_total = db_session().query(func.count(Order.id)).scalar()
        orders_pending = db_session().query(func.count(Order.id)).filter(
            Order.payment_status.in_(["awaiting_transfer", "receipt_uploaded"])
        ).scalar()
        orders_confirmed = db_session().query(func.count(Order.id)).filter(
            Order.payment_status == "confirmed"
        ).scalar()
        orders_pod = db_session().query(func.count(Order.id)).filter(
            Order.payment_method == "pay_on_delivery"
        ).scalar()

        # Receipts awaiting admin review
        receipts_pending = db_session().query(func.count(Payment.id)).filter(
            Payment.status == "receipt_uploaded"
        ).scalar()

        # Users
        users_total = db_session().query(func.count(User.id)).scalar()
        users_flagged = db_session().query(func.count(User.id)).filter(
            User.needs_monitoring.is_(True)
        ).scalar()

        # Products
        products_active = db_session().query(func.count(Product.id)).filter(
            Product.active.is_(True)
        ).scalar()
        products_low_stock = db_session().query(func.count(Product.id)).filter(
            Product.active.is_(True),
            Product.stock > 0,
            Product.stock <= 3,
        ).scalar()
        products_out_of_stock = db_session().query(func.count(Product.id)).filter(
            Product.active.is_(True),
            Product.stock == 0,
        ).scalar()

        # Installations
        installations_total = db_session().query(func.count(Installation.id)).scalar()
        installations_pending = db_session().query(func.count(Installation.id)).filter(
            Installation.status == "pending"
        ).scalar()

        # Blog
        blog_posts_published = db_session().query(func.count(BlogPost.id)).filter(
            BlogPost.is_published.is_(True)
        ).scalar()
        blog_subscribers = db_session().query(func.count(BlogSubscriber.id)).filter(
            BlogSubscriber.is_active.is_(True)
        ).scalar()

        # Jobs
        jobs_active = db_session().query(func.count(JobListing.id)).filter(
            JobListing.active.is_(True)
        ).scalar()
        jobs_featured = db_session().query(func.count(JobListing.id)).filter(
            JobListing.active.is_(True),
            JobListing.featured.is_(True),
        ).scalar()

        # Feedback
        feedback_new = db_session().query(func.count(Feedback.id)).filter(
            Feedback.status == "new"
        ).scalar()
        feedback_total = db_session().query(func.count(Feedback.id)).scalar()

        payload = {
            # Revenue
            "revenue": float(confirmed_revenue or 0),
            "revenue_total": float(total_revenue or 0),
            "currency": settings.default_currency,
            # Orders
            "orders": orders_total,
            "orders_pending": orders_pending,
            "orders_confirmed": orders_confirmed,
            "orders_pod": orders_pod,
            "receipts_pending": receipts_pending,
            # Users
            "users": users_total,
            "users_flagged": users_flagged,
            # Products
            "products": products_active,
            "products_low_stock": products_low_stock,
            "products_out_of_stock": products_out_of_stock,
            # Installations
            "installations": installations_total,
            "installations_pending": installations_pending,
            # Blog
            "blog_posts": blog_posts_published,
            "blog_subscribers": blog_subscribers,
            # Jobs
            "jobs": jobs_active,
            "jobs_featured": jobs_featured,
            # Feedback
            "feedback_new": feedback_new,
            "feedback_total": feedback_total,
        }
        return json_success(payload)

    @app.get("/api/admin/reports")
    def admin_reports():
        authenticate(admin=True)
        return json_success(build_admin_report_payload())

    @app.get("/api/admin/reports/export")
    def export_admin_reports():
        authenticate(admin=True)
        export_format = (request.args.get("format") or "xlsx").strip().lower()
        report = build_admin_report_payload()

        if export_format == "pdf":
            return Response(
                admin_report_pdf(report),
                mimetype="application/pdf",
                headers={
                    "Content-Disposition": 'attachment; filename="hoinam-admin-report.pdf"'
                },
            )

        if export_format not in {"xlsx", "excel"}:
            raise ApiError("Supported report export formats are xlsx and pdf.", 400)

        from openpyxl import Workbook

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        summary_sheet.append(["Metric", "Value"])
        for key, value in report["summary"].items():
            summary_sheet.append([key.replace("_", " ").title(), value])
        summary_sheet.append([])
        for key, value in report["inventory_status"].items():
            summary_sheet.append([key.replace("_", " ").title(), value])

        sales_sheet = workbook.create_sheet("Sales History")
        sales_sheet.append(["Order", "Customer", "Status", "Payment", "Total", "Date"])
        for order in report["sales_history"]:
            customer = order.get("user") or {}
            sales_sheet.append(
                [
                    order.get("order_number"),
                    customer.get("full_name") or customer.get("email") or "Customer",
                    order.get("status"),
                    order.get("payment_status"),
                    order.get("total_amount"),
                    order.get("created_at"),
                ]
            )

        inventory_sheet = workbook.create_sheet("Inventory")
        inventory_sheet.append(["Product", "SKU", "Stock", "Price", "Category", "Brand"])
        for product in report["inventory_products"]:
            inventory_sheet.append(
                [
                    product.get("name"),
                    product.get("sku"),
                    product.get("stock"),
                    product.get("price"),
                    product.get("category"),
                    product.get("brand"),
                ]
            )

        movement_sheet = workbook.create_sheet("Stock Movements")
        movement_sheet.append(["Date", "Product", "Type", "Quantity", "Previous", "New", "Order", "Note"])
        for movement in report["stock_movements"]:
            movement_sheet.append(
                [
                    movement.get("created_at"),
                    movement.get("product_name"),
                    movement.get("movement_type"),
                    movement.get("quantity"),
                    movement.get("previous_stock"),
                    movement.get("new_stock"),
                    movement.get("order_number"),
                    movement.get("note"),
                ]
            )

        performance_sheet = workbook.create_sheet("Product Performance")
        performance_sheet.append(["Product", "Quantity Sold", "Gross Revenue", "Confirmed Revenue", "Order Lines"])
        for item in report["product_performance"]:
            performance_sheet.append(
                [
                    item.get("product_name"),
                    item.get("quantity_sold"),
                    item.get("gross_revenue"),
                    item.get("confirmed_revenue"),
                    item.get("orders"),
                ]
            )

        enquiries_sheet = workbook.create_sheet("Enquiries")
        enquiries_sheet.append(["Date", "Name", "Email", "Phone", "Type", "Status", "Message"])
        for item in report["customer_enquiries"]:
            enquiries_sheet.append(
                [
                    item.get("created_at"),
                    item.get("name"),
                    item.get("email"),
                    item.get("phone"),
                    item.get("service_type"),
                    item.get("status"),
                    item.get("message"),
                ]
            )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="hoinam-admin-report.xlsx"'
            },
        )

    @app.post("/api/admin/upload-inventory")
    def upload_inventory():
        user = authenticate(admin=True)
        file_storage = request.files.get("file")
        if not file_storage:
            raise ApiError("Please attach an Excel file.", 400)

        rows = parse_inventory_file(file_storage)
        if not rows:
            raise ApiError(
                "No inventory rows were found in the uploaded workbook.", 400
            )

        created = 0
        updated = 0
        for row in rows:
            product_slug = row.get("slug") or slugify(row["name"])
            product_sku = row.get("sku") or product_slug.upper().replace("-", "_")
            legacy_slug = row.get("legacy_slug")
            product = find_existing_inventory_product(
                row, product_sku, product_slug, legacy_slug
            )
            if product is None:
                product = Product(
                    name=row["name"],
                    slug=product_slug,
                    sku=product_sku,
                    brand=row.get("brand"),
                    store_slug=row.get("store_slug"),
                    category=row["category"],
                    summary=row["summary"],
                    description=row["description"],
                    price=to_decimal(row["price"]),
                    currency=settings.default_currency,
                    stock=row["stock"],
                    image_url=resolve_product_image_url(
                        row["name"], row["image_url"], product_slug
                    ),
                    active=True,
                )
                db_session().add(product)
                db_session().flush()
                record_inventory_movement(
                    product=product,
                    movement_type="import",
                    quantity=int(product.stock or 0),
                    previous_stock=0,
                    new_stock=int(product.stock or 0),
                    actor=user,
                    note="Created from Excel inventory upload.",
                )
                created += 1
            else:
                previous_stock = int(product.stock or 0)
                product.name = row["name"]
                product.slug = product_slug
                product.sku = product_sku
                product.brand = row.get("brand") or product.brand
                product.store_slug = row.get("store_slug") or product.store_slug
                product.stock = row["stock"]
                if row["price"]:
                    product.price = to_decimal(row["price"])
                product.category = row["category"] or product.category
                product.image_url = resolve_product_image_url(
                    row["name"],
                    row["image_url"] or product.image_url,
                    product.slug or product_slug,
                )
                product.summary = row["summary"] or product.summary
                product.description = row["description"] or product.description
                new_stock = int(product.stock or 0)
                if previous_stock != new_stock:
                    record_inventory_movement(
                        product=product,
                        movement_type="import",
                        quantity=new_stock - previous_stock,
                        previous_stock=previous_stock,
                        new_stock=new_stock,
                        actor=user,
                        note="Stock synced from Excel inventory upload.",
                    )
                updated += 1

        db_session().commit()
        return json_success(
            {"created": created, "updated": updated},
            message="Inventory upload complete.",
        )

    @app.get("/api/admin/pending-deliveries")
    def get_pending_deliveries():
        """Get orders awaiting delivery confirmation (admin only)."""
        authenticate(admin=True)
        pending_orders = (
            db_session()
            .query(Order)
            .filter(Order.status.in_(["payment_pending", "confirmed"]))
            .options(joinedload(Order.user))
            .order_by(Order.created_at.asc())
            .all()
        )
        payload = []
        for order in pending_orders:
            data = order.to_dict()
            data["user"] = order.user.to_dict() if order.user else None
            payload.append(data)
        return json_success(payload)

    @app.post("/api/admin/orders/<int:order_id>/confirm-delivery")
    def confirm_delivery(order_id: int):
        """Mark order as approved/delivered and email the customer (admin only)."""
        authenticate(admin=True)

        order = (
            db_session()
            .query(Order)
            .options(joinedload(Order.user))
            .filter(Order.id == order_id)
            .first()
        )
        if not order:
            raise ApiError("Order not found.", 404)

        order.status = "delivered"
        order.payment_status = "confirmed"

        payment = (
            db_session().query(Payment).filter(Payment.order_id == order_id).first()
        )
        if payment:
            payment.status = "confirmed"
            payment.notes = (
                f"Payment confirmed by admin on {datetime.now(timezone.utc).isoformat()}."
            )

        ensure_order_documents(order)
        db_session().commit()

        # Email the customer their approval notification
        if smtp_is_configured(settings) and order.user and order.user.email:
            try:
                approval_msg = build_order_approved_message(
                    settings=settings,
                    user=order.user,
                    order=order,
                    shipping_address=order.shipping_address,
                )
                send_message_via_smtp(settings, approval_msg)
            except Exception:
                app.logger.exception(
                    "Order %s approved but approval email could not be sent.",
                    order.order_number,
                )

        return json_success(
            order.to_dict(),
            message=f"Order {order.order_number} approved. Customer notified by email.",
        )

    # ── Seasonal UI ───────────────────────────────────────────────────────────

    @app.get("/api/season")
    def get_season():
        """Get the current active season theme (public)."""
        # Stored in a simple in-memory + file-backed config
        import json as _json
        season_file = project_root / ".season.json"
        if season_file.is_file():
            try:
                data = _json.loads(season_file.read_text())
                return json_success(data)
            except Exception:
                pass
        return json_success({"season": "default", "banner": "", "active": False})

    @app.post("/api/admin/season")
    def set_season():
        """Set the active season theme (admin only)."""
        import json as _json
        authenticate(admin=True)
        payload = request.get_json(silent=True) or {}
        allowed_seasons = {
            "default", "christmas", "new_year", "easter", "eid",
            "independence", "valentine", "halloween", "custom"
        }
        season = (payload.get("season") or "default").strip().lower()
        if season not in allowed_seasons:
            season = "default"
        data = {
            "season": season,
            "banner": (payload.get("banner") or "").strip(),
            "active": bool(payload.get("active", season != "default")),
        }
        season_file = project_root / ".season.json"
        season_file.write_text(_json.dumps(data))
        return json_success(data, message="Season theme updated.")

    # ── Coupons ───────────────────────────────────────────────────────────────

    @app.post("/api/coupons/validate")
    def validate_coupon():
        """Validate a coupon code against a subtotal (public — no auth required)."""
        payload = request.get_json(silent=True) or {}
        code = (payload.get("code") or "").strip().upper()
        subtotal = to_decimal(payload.get("subtotal") or 0)

        if not code:
            raise ApiError("Coupon code is required.", 400)

        coupon = (
            db_session()
            .query(Coupon)
            .filter(Coupon.code == code, Coupon.is_active.is_(True))
            .first()
        )

        if not coupon:
            raise ApiError("Coupon code not found or is no longer active.", 404)
        if coupon.expires_at and coupon.expires_at < datetime.now(timezone.utc):
            raise ApiError("This coupon has expired.", 400)
        if coupon.max_uses is not None and coupon.uses >= coupon.max_uses:
            raise ApiError("This coupon has reached its usage limit.", 400)
        if subtotal > 0 and subtotal < coupon.min_order_amount:
            raise ApiError(
                f"This coupon requires a minimum order of "
                f"{float(coupon.min_order_amount):,.2f} {settings.default_currency}.",
                400,
            )

        if coupon.discount_type == "percent":
            discount = (subtotal * coupon.discount_value / Decimal("100")).quantize(Decimal("0.01"))
        else:
            discount = min(coupon.discount_value, subtotal)

        return json_success({
            "code": coupon.code,
            "description": coupon.description,
            "discount_type": coupon.discount_type,
            "discount_value": float(coupon.discount_value),
            "discount_amount": float(discount),
            "final_total": float(max(subtotal - discount, Decimal("0.00"))),
            "currency": settings.default_currency,
        })

    @app.get("/api/admin/coupons")
    def list_coupons():
        authenticate(admin=True)
        coupons = (
            db_session()
            .query(Coupon)
            .order_by(Coupon.created_at.desc())
            .all()
        )
        return json_success([c.to_dict() for c in coupons])

    @app.post("/api/admin/coupons")
    def create_coupon():
        authenticate(admin=True)
        payload = request.get_json(silent=True) or {}

        code = (payload.get("code") or "").strip().upper()
        if not code:
            raise ApiError("Coupon code is required.", 400)

        existing = db_session().query(Coupon).filter(Coupon.code == code).first()
        if existing:
            raise ApiError("A coupon with this code already exists.", 409)

        discount_type = (payload.get("discount_type") or "percent").strip()
        if discount_type not in {"percent", "fixed"}:
            raise ApiError("discount_type must be 'percent' or 'fixed'.", 400)

        discount_value = to_decimal(payload.get("discount_value") or 0)
        if discount_value <= 0:
            raise ApiError("Discount value must be greater than zero.", 400)
        if discount_type == "percent" and discount_value > 100:
            raise ApiError("Percentage discount cannot exceed 100.", 400)

        expires_at = None
        if payload.get("expires_at"):
            try:
                expires_at = datetime.fromisoformat(payload["expires_at"].replace("Z", "+00:00"))
            except ValueError:
                raise ApiError("Invalid expires_at date format. Use ISO 8601.", 400)

        coupon = Coupon(
            code=code,
            description=(payload.get("description") or "").strip() or None,
            discount_type=discount_type,
            discount_value=discount_value,
            min_order_amount=to_decimal(payload.get("min_order_amount") or 0),
            max_uses=int(payload["max_uses"]) if payload.get("max_uses") else None,
            is_active=bool(payload.get("is_active", True)),
            expires_at=expires_at,
        )
        db_session().add(coupon)
        db_session().commit()
        return json_success(coupon.to_dict(), status_code=201, message="Coupon created.")

    @app.put("/api/admin/coupons/<int:coupon_id>")
    def update_coupon(coupon_id: int):
        authenticate(admin=True)
        coupon = db_session().query(Coupon).filter(Coupon.id == coupon_id).first()
        if not coupon:
            raise ApiError("Coupon not found.", 404)

        payload = request.get_json(silent=True) or {}
        if "description" in payload:
            coupon.description = (payload.get("description") or "").strip() or None
        if "discount_type" in payload:
            dt = (payload.get("discount_type") or "percent").strip()
            if dt not in {"percent", "fixed"}:
                raise ApiError("discount_type must be 'percent' or 'fixed'.", 400)
            coupon.discount_type = dt
        if "discount_value" in payload:
            dv = to_decimal(payload.get("discount_value") or 0)
            if dv <= 0:
                raise ApiError("Discount value must be greater than zero.", 400)
            coupon.discount_value = dv
        if "min_order_amount" in payload:
            coupon.min_order_amount = to_decimal(payload.get("min_order_amount") or 0)
        if "max_uses" in payload:
            coupon.max_uses = int(payload["max_uses"]) if payload.get("max_uses") else None
        if "is_active" in payload:
            coupon.is_active = bool(payload.get("is_active"))
        if "expires_at" in payload:
            if payload["expires_at"]:
                try:
                    coupon.expires_at = datetime.fromisoformat(
                        payload["expires_at"].replace("Z", "+00:00")
                    )
                except ValueError:
                    raise ApiError("Invalid expires_at date format.", 400)
            else:
                coupon.expires_at = None

        db_session().commit()
        return json_success(coupon.to_dict(), message="Coupon updated.")

    @app.delete("/api/admin/coupons/<int:coupon_id>")
    def delete_coupon(coupon_id: int):
        authenticate(admin=True)
        coupon = db_session().query(Coupon).filter(Coupon.id == coupon_id).first()
        if not coupon:
            raise ApiError("Coupon not found.", 404)
        db_session().delete(coupon)
        db_session().commit()
        return json_success({"id": coupon_id}, message="Coupon deleted.")

    # ── Feedback ──────────────────────────────────────────────────────────────

    @app.post("/api/feedback")
    def submit_feedback():
        """Submit customer feedback (public — no auth required)."""
        payload = request.get_json(silent=True) or {}

        name = (payload.get("name") or "").strip()
        if not name:
            raise ApiError("Your name is required.", 400)

        message_text = (payload.get("message") or "").strip()
        if not message_text:
            raise ApiError("A feedback message is required.", 400)

        email = (payload.get("email") or "").strip() or None
        phone = (payload.get("phone") or "").strip() or None
        service_type = (payload.get("service_type") or "general").strip()
        allowed_types = {"general", "pre_service", "post_service", "product", "installation"}
        if service_type not in allowed_types:
            service_type = "general"

        raw_rating = payload.get("rating")
        rating = None
        if raw_rating is not None:
            try:
                rating = int(raw_rating)
                if not 1 <= rating <= 5:
                    rating = None
            except (ValueError, TypeError):
                rating = None

        feedback = Feedback(
            name=name,
            email=email,
            phone=phone,
            service_type=service_type,
            rating=rating,
            message=message_text,
            order_number=(payload.get("order_number") or "").strip() or None,
            status="new",
        )
        db_session().add(feedback)
        db_session().commit()

        if smtp_is_configured(settings):
            # Notify support team
            try:
                notif = build_feedback_notification_message(
                    settings=settings,
                    feedback=feedback,
                )
                send_message_via_smtp(settings, notif)
            except Exception:
                app.logger.exception("Feedback notification email failed for feedback %s.", feedback.id)

            # Send acknowledgement to customer if they gave an email
            if feedback.email:
                try:
                    ack = build_feedback_acknowledgement_message(
                        settings=settings,
                        feedback=feedback,
                    )
                    send_message_via_smtp(settings, ack)
                except Exception:
                    app.logger.exception("Feedback acknowledgement email failed for %s.", feedback.email)

        return json_success(
            feedback.to_dict(),
            status_code=201,
            message="Thank you! Your feedback has been received.",
        )

    @app.get("/api/admin/feedback")
    def list_feedback():
        """List all feedback submissions (admin only)."""
        authenticate(admin=True)
        items = (
            db_session()
            .query(Feedback)
            .order_by(Feedback.created_at.desc())
            .all()
        )
        return json_success([f.to_dict() for f in items])

    @app.put("/api/admin/feedback/<int:feedback_id>")
    def update_feedback_status(feedback_id: int):
        """Update feedback status (admin only)."""
        authenticate(admin=True)
        item = db_session().query(Feedback).filter(Feedback.id == feedback_id).first()
        if not item:
            raise ApiError("Feedback not found.", 404)
        payload = request.get_json(silent=True) or {}
        if "status" in payload:
            item.status = payload["status"]
        db_session().commit()
        return json_success(item.to_dict())

    @app.get("/")
    def serve_index():
        return send_from_directory(project_root, "index.html")

    @app.get("/<path:path>")
    def serve_frontend(path: str):
        # Don't intercept API routes — let Flask handle them
        if path.startswith("api/"):
            raise ApiError("Not found.", 404)

        candidate = project_root / path
        if candidate.is_file():
            return send_from_directory(project_root, path)

        if not candidate.suffix:
            html_candidate = project_root / f"{path}.html"
            if html_candidate.is_file():
                return send_from_directory(project_root, f"{path}.html")

            index_candidate = project_root / path / "index.html"
            if index_candidate.is_file():
                return send_from_directory(project_root, f"{path}/index.html")

        # Serve the custom 404 page for unknown frontend routes
        four_oh_four = project_root / "404.html"
        if four_oh_four.is_file():
            return send_from_directory(project_root, "404.html"), 404
        raise ApiError("Page not found.", 404)

    return app
