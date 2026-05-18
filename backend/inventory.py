from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .utils import resolve_product_image_url, slugify, to_decimal


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INVENTORY_PATH = ROOT / "STOCK INVENTORY.xlsx"

BRAND_NAMES = {
    "ECOFLOW": "EcoFlow",
    "BLUETTI": "Bluetti",
    "DEYE": "Deye",
}

STORE_SLUGS = {
    "ECOFLOW": "ecoflow",
    "BLUETTI": "bluetti",
    "DEYE": "deye",
}

FEATURED_REFERENCES = {
    "river 2",
    "river 2 max",
    "river 2 pro",
    "delta 2 max",
    "delta pro",
    "delta pro 3",
    "bluetti ac180",
    "bluetti ac200pl",
    "bluetti ac300",
    "bluetti ac500",
    "bluetti ep500p",
    "sun-6k-sg04lp1-eu-sm2",
    "sun-12k-sg04lp3-eu-am3",
    "se-g5.1 5kwh 51.2v",
}

def clean_text(value) -> str:
    return " ".join(str(value or "").replace("_", " ").split())


def is_brand_header_row(
    sn: str | None, row_name: str | None, reference: str | None, description: str | None
) -> bool:
    return bool(clean_text(sn)) and not any(
        [
            clean_text(row_name),
            clean_text(reference),
            clean_text(description),
        ]
    )


def display_brand(raw_brand: str | None) -> str:
    key = clean_text(raw_brand).upper()
    return BRAND_NAMES.get(key, clean_text(raw_brand) or "Hoinam")


def store_slug(raw_brand: str | None) -> str:
    key = clean_text(raw_brand).upper()
    return STORE_SLUGS.get(key, slugify(display_brand(raw_brand)))


def product_display_name(brand: str, reference: str) -> str:
    reference = clean_text(reference)
    if not reference:
        return brand
    if reference.lower().startswith(brand.lower()):
        return reference
    return f"{brand} {reference}"


def trim_summary(value: str, max_length: int = 180) -> str:
    text = clean_text(value)
    if len(text) <= max_length:
        return text
    truncated = text[: max_length - 1].rsplit(" ", 1)[0]
    return f"{truncated}..."


def product_summary(name: str, description: str, category: str) -> str:
    if description:
        return trim_summary(description)
    return f"{name} in the Hoinam Energy {category.lower()} catalog."


def product_highlights(category: str, brand: str, description: str) -> list[str]:
    category_key = clean_text(category).lower()
    haystack = f"{category_key} {description}".lower()
    highlights: list[str] = []

    if category_key == "portable power":
        highlights.append("Portable backup power")
    elif category_key == "solar panels":
        highlights.append("Solar charging ready")
    elif category_key == "inverters":
        highlights.append("Hybrid inverter solution")
    elif category_key == "batteries":
        highlights.append("Energy storage expansion")
    elif category_key == "accessories":
        highlights.append("System accessory")
    else:
        highlights.append("Energy backup solution")

    if "solar" in haystack:
        highlights.append("Solar compatible")
    if any(term in haystack for term in ["battery", "kwh", "storage"]):
        highlights.append("Battery storage support")
    if any(term in haystack for term in ["portable", "river", "delta", "ac"]):
        highlights.append("Designed for flexible backup")

    highlights.append(f"{brand} product support")
    return list(dict.fromkeys(highlights))[:4]


def is_featured_product(reference: str, stock: int) -> bool:
    return stock > 0 and clean_text(reference).lower() in FEATURED_REFERENCES


def infer_category(row_name: str | None, reference: str, description: str, brand: str) -> str:
    label = clean_text(row_name).lower()
    haystack = f"{label} {reference} {description}".lower()

    if "solar panel" in haystack:
        return "Solar Panels"
    if "portable power station" in label or any(term in haystack for term in ["river", "delta", "ac180", "eb3a", "ep500"]):
        return "Portable Power"
    if "inverter" in haystack or "sun-" in haystack:
        return "Inverters"
    if any(term in haystack for term in ["battery", "b300", "b300s", "b700", "kwh", "bos-"]):
        return "Batteries"
    if any(term in haystack for term in ["meter", "junction", "base", "trolly", "trolley", "pdu"]):
        return "Accessories"
    if "+" in reference:
        return "Energy Kits"
    if brand == "Deye":
        return "Inverters"
    return "Energy Systems"


def parse_stock_inventory(path: str | Path | None = None) -> list[dict]:
    source = path or DEFAULT_INVENTORY_PATH
    if hasattr(source, "read"):
        workbook = load_workbook(source, data_only=True)
    else:
        inventory_path = Path(source)
        if not inventory_path.is_file():
            return []
        workbook = load_workbook(inventory_path, data_only=True)
    sheet = workbook.active
    current_brand = None
    products: list[dict] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        sn, row_name, reference, description, quantity, rate = (list(row) + [None] * 6)[:6]

        row_name_text = clean_text(row_name)
        reference_text = clean_text(reference)
        description_text = clean_text(description)

        if is_brand_header_row(sn, row_name_text, reference_text, description_text):
            current_brand = clean_text(sn)
            continue

        if not reference_text:
            continue

        brand = display_brand(current_brand)
        brand_slug = store_slug(current_brand)
        name = product_display_name(brand, reference_text)
        slug = slugify(name)
        legacy_slug = slugify(reference_text)
        category = infer_category(row_name_text, reference_text, description_text, brand)
        stock = int(quantity) if quantity not in (None, "") else 0
        price = to_decimal(rate)
        image_url = (
            resolve_product_image_url(name, None, slug)
            or resolve_product_image_url(reference_text, None, legacy_slug)
        )
        summary = product_summary(name, description_text, category)

        products.append(
            {
                "name": name,
                "slug": slug,
                "sku": f"{brand_slug.upper()}-{legacy_slug.upper().replace('-', '_')}",
                "brand": brand,
                "store_slug": brand_slug,
                "category": category,
                "summary": summary,
                "description": description_text or f"{name} imported from the Hoinam Energy stock inventory.",
                "price": price,
                "currency": "NGN",
                "stock": stock,
                "image_url": image_url,
                "highlights": product_highlights(category, brand, description_text),
                "featured": is_featured_product(reference_text, stock),
                "active": True,
                "specs": {
                    "reference": reference_text,
                    "inventory_name": row_name_text,
                },
                "legacy_slug": legacy_slug,
                "reference": reference_text,
            }
        )

    return products
