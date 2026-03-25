import os

from openpyxl import load_workbook

from app import create_app, db
from app.models import Product


def load_inventory(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    header = [str(cell.value or '').strip().lower() for cell in sheet[1]]

    added = 0
    updated = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {header[idx]: value for idx, value in enumerate(row) if idx < len(header)}
        name = row_data.get('product name') or row_data.get('name')
        quantity = row_data.get('quantity') or 0
        price = row_data.get('price') or 0
        description = row_data.get('description')
        category = row_data.get('category')
        image_url = row_data.get('image_url')

        if not name:
            continue

        product = Product.query.filter_by(name=name).first()
        if product:
            product.quantity = int(quantity)
            product.price = price
            product.description = description or product.description
            product.category = category or product.category
            product.image_url = image_url or product.image_url
            updated += 1
        else:
            product = Product(
                name=name,
                quantity=int(quantity),
                price=price,
                description=description,
                category=category,
                image_url=image_url,
            )
            db.session.add(product)
            added += 1

    db.session.commit()
    return added, updated


if __name__ == '__main__':
    app = create_app()
    with app.app_context():
        inventory_file = os.path.join(os.path.dirname(__file__), '..', '..', 'inventory.xlsx')
        added, updated = load_inventory(inventory_file)
        print(f"Inventory sync complete. Added: {added}, Updated: {updated}")
