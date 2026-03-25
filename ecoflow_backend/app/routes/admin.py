import io

from flask import Blueprint, jsonify, request
from openpyxl import load_workbook

from app import db
from app.models import Installation, Order, Product, User
from app.routes.helpers import admin_required


admin_bp = Blueprint('admin', __name__)


def _serialize_user(user):
    return {
        "id": user.id,
        "name": user.name,
        "username": user.username,
        "email": user.email,
        "phone": user.phone,
        "role": user.role,
        "created_at": user.created_at.isoformat() if user.created_at else None,
    }


def _serialize_order(order):
    return {
        "id": order.id,
        "user_id": order.user_id,
        "items": order.items,
        "total_amount": float(order.total_amount or 0),
        "status": order.status,
        "payment_status": order.payment_status,
        "created_at": order.created_at.isoformat() if order.created_at else None,
    }


def _serialize_installation(installation):
    return {
        "id": installation.id,
        "user_id": installation.user_id,
        "preferred_date": installation.preferred_date,
        "address": installation.address,
        "status": installation.status,
        "created_at": installation.created_at.isoformat() if installation.created_at else None,
    }


@admin_bp.get('/stats')
@admin_required

def admin_stats():
    return jsonify(
        {
            "users": User.query.count(),
            "products": Product.query.count(),
            "orders": Order.query.count(),
            "installations": Installation.query.count(),
        }
    )


@admin_bp.get('/users')
@admin_required

def list_users():
    users = User.query.order_by(User.created_at.desc()).all()
    return jsonify([_serialize_user(user) for user in users])


@admin_bp.get('/orders')
@admin_required

def list_orders():
    orders = Order.query.order_by(Order.created_at.desc()).all()
    return jsonify([_serialize_order(order) for order in orders])


@admin_bp.get('/installations')
@admin_required

def list_installations():
    installations = Installation.query.order_by(Installation.created_at.desc()).all()
    return jsonify([_serialize_installation(item) for item in installations])


@admin_bp.post('/upload-inventory')
@admin_required

def upload_inventory():
    if 'file' not in request.files:
        return jsonify({"error": "Excel file is required"}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({"error": "File name is required"}), 400

    in_memory = io.BytesIO(file.read())
    workbook = load_workbook(in_memory)
    sheet = workbook.active

    header = [str(cell.value or '').strip().lower() for cell in sheet[1]]
    rows_added = 0
    rows_updated = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {header[idx]: value for idx, value in enumerate(row) if idx < len(header)}
        name = row_data.get('product name') or row_data.get('name')
        quantity = row_data.get('quantity') or 0
        price = row_data.get('price') or 0
        description = row_data.get('description')
        category = row_data.get('category')
        image_url = row_data.get('image_url')

        if not name:
            continue

        product = Product.query.filter_by(name=name).first()
        if product:
            product.quantity = int(quantity)
            product.price = price
            product.description = description or product.description
            product.category = category or product.category
            product.image_url = image_url or product.image_url
            rows_updated += 1
        else:
            product = Product(
                name=name,
                quantity=int(quantity),
                price=price,
                description=description,
                category=category,
                image_url=image_url,
            )
            db.session.add(product)
            rows_added += 1

    db.session.commit()

    return jsonify({"added": rows_added, "updated": rows_updated})
